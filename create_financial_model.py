#!/usr/bin/env python3
"""
VoyageRent - Car Rental Financial Model Generator
Creates a professional Excel financial model for a car rental business in Spain.
All calculations flow from the INPUTS sheet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

wb = openpyxl.Workbook()

# ── Style definitions ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="1B3A5C", size=11)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FONT = Font(name="Calibri", size=11, color="000000")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RESULT_FONT = Font(name="Calibri", bold=True, size=11, color="1B3A5C")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1B3A5C")
KPI_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
NEGATIVE_FONT = Font(name="Calibri", bold=True, size=11, color="FF0000")
POSITIVE_FONT = Font(name="Calibri", bold=True, size=11, color="006100")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

EUR_FORMAT = '#,##0.00 €'
EUR_FORMAT_0 = '#,##0 €'
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'
MONTHS_FORMAT = '0.0'


def style_header(ws, row, max_col, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, max_col, fill=None, font=NORMAL_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.font = font
        cell.border = thin_border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: INPUTS
# ═══════════════════════════════════════════════════════════════════════════
ws_inputs = wb.active
ws_inputs.title = "INPUTS"
ws_inputs.sheet_properties.tabColor = "1B3A5C"

set_col_widths(ws_inputs, [5, 45, 18, 20, 35])

# Title
ws_inputs.merge_cells('B1:D1')
ws_inputs['B1'] = "VOYAGERENT — ОСНОВНЫЕ ПАРАМЕТРЫ (INPUTS)"
ws_inputs['B1'].font = TITLE_FONT

ws_inputs.merge_cells('B2:D2')
ws_inputs['B2'] = "Все данные вводятся ТОЛЬКО на этом листе. Остальные листы рассчитываются автоматически."
ws_inputs['B2'].font = Font(name="Calibri", italic=True, size=10, color="666666")

# Helper: write input sections
def write_input_section(ws, start_row, title, items):
    """items: list of (label, value, unit, comment)"""
    ws.cell(row=start_row, column=2, value=title)
    style_row(ws, start_row, 5, SECTION_FILL, SECTION_FONT)
    r = start_row + 1
    for label, value, unit, comment in items:
        ws.cell(row=r, column=2, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = thin_border
        c = ws.cell(row=r, column=3, value=value)
        c.fill = INPUT_FILL
        c.font = INPUT_FONT
        c.border = thin_border
        if isinstance(value, float) and value < 1 and value > 0:
            c.number_format = PCT_FORMAT
        elif isinstance(value, (int, float)) and abs(value) >= 100:
            c.number_format = EUR_FORMAT_0 if unit in ['€', '€/мес', '€/год', '€/авто/мес'] else NUM_FORMAT
        ws.cell(row=r, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=5, value=comment).font = Font(name="Calibri", size=9, italic=True, color="888888")
        ws.cell(row=r, column=5).border = thin_border
        r += 1
    return r + 1

# Column headers
for col, val in [(2, "Параметр"), (3, "Значение"), (4, "Единица"), (5, "Комментарий")]:
    ws_inputs.cell(row=3, column=col, value=val)
style_header(ws_inputs, 3, 5)

# Section 1: Fleet
r = write_input_section(ws_inputs, 5, "🚗 АВТОПАРК (Fleet)", [
    ("Количество автомобилей", 100, "шт.", "Текущий размер автопарка"),           # C6
    ("Стоимость автомобиля (средняя)", 18000, "€", "Средняя закупочная цена"),      # C7
    ("Срок амортизации", 5, "лет", "Период списания стоимости"),                    # C8
    ("Остаточная стоимость (%)", 0.20, "%", "% от стоимости при продаже через срок амортизации"),  # C9
])

# Section 2: Revenue
r = write_input_section(ws_inputs, r, "💰 ДОХОДЫ (Revenue)", [
    ("Средняя цена аренды в день", 45, "€/день", "Средневзвешенная по классам авто"),   # C12
    ("Средняя продолжительность аренды", 5, "дней", "Средний срок одной аренды"),        # C13
    ("Загрузка автопарка", 0.65, "%", "% времени, когда авто сдано в аренду"),           # C14
    ("Количество аренд в месяц (на 1 авто)", 4, "шт.", "Среднее кол-во аренд на 1 авто в месяц"),  # C15
    ("Комиссия агрегаторов", 0.15, "%", "Booking, Rentalcars, Discovercars и т.д."),    # C16
])

# Section 3: Operating costs per car
r = write_input_section(ws_inputs, r, "🔧 РАСХОДЫ НА АВТОМОБИЛЬ (Per Vehicle Costs)", [
    ("Страховка (на авто в месяц)", 120, "€/авто/мес", "Полная страховка + КАСКО"),    # C19
    ("Ремонт и обслуживание (на авто в месяц)", 80, "€/авто/мес", "Средний месячный расход"),  # C20
    ("Парковка (на авто в месяц)", 50, "€/авто/мес", "Стоянка / паркинг"),             # C21
    ("Топливо / подготовка (на авто в месяц)", 40, "€/авто/мес", "Подготовка, мойка, топливо"),  # C22
    ("ITV + налоги (на авто в год)", 200, "€/авто/год", "Техосмотр + транспортный налог"),  # C23
])

# Section 4: Company overhead
r = write_input_section(ws_inputs, r, "🏢 ОБЩИЕ РАСХОДЫ КОМПАНИИ (Overhead)", [
    ("Количество сотрудников", 6, "чел.", "Штат компании"),                          # C26
    ("Средняя зарплата (gross)", 1800, "€/мес", "Зарплата до налогов"),               # C27
    ("Налоги на зарплату (%)", 0.30, "%", "Соц. взносы работодателя в Испании"),      # C28
    ("Аренда офиса / точки", 1500, "€/мес", "Офис + точки выдачи"),                   # C29
    ("Маркетинг и реклама", 2000, "€/мес", "Google Ads, SEO, соц. сети"),             # C30
    ("Софт и IT", 500, "€/мес", "CRM, GPS-трекинг, бухгалтерия"),                     # C31
    ("Связь и интернет", 200, "€/мес", "Телефония, интернет"),                        # C32
    ("Прочие расходы", 500, "€/мес", "Непредвиденные расходы"),                       # C33
])

# Section 5: Growth & Scenarios
r = write_input_section(ws_inputs, r, "📈 РОСТ И СЦЕНАРИИ (Growth)", [
    ("Планируемый рост флота (% в год)", 0.20, "%", "На сколько % вырастет флот за год"),  # C36
    ("Сезонность: высокий сезон (мес.)", 5, "мес.", "Кол-во месяцев высокого сезона"),    # C37
    ("Коэффициент высокого сезона", 1.30, "×", "Множитель загрузки в высокий сезон"),     # C38
    ("Коэффициент низкого сезона", 0.70, "×", "Множитель загрузки в низкий сезон"),       # C39
])

# Section 6: Tax
r = write_input_section(ws_inputs, r, "🏛️ НАЛОГИ (Tax)", [
    ("Ставка налога на прибыль", 0.25, "%", "Impuesto de Sociedades (Испания)"),  # C42
    ("НДС (IVA)", 0.21, "%", "IVA в Испании 21%"),                               # C43
])


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: UNIT ECONOMICS
# ═══════════════════════════════════════════════════════════════════════════
ws_unit = wb.create_sheet("UNIT ECONOMICS")
ws_unit.sheet_properties.tabColor = "2E75B6"
set_col_widths(ws_unit, [5, 45, 20, 15, 35])

ws_unit.merge_cells('B1:D1')
ws_unit['B1'] = "UNIT ECONOMICS — Экономика одного автомобиля"
ws_unit['B1'].font = TITLE_FONT

for col, val in [(2, "Показатель"), (3, "Значение"), (4, "Ед."), (5, "Формула / Комментарий")]:
    ws_unit.cell(row=3, column=col, value=val)
style_header(ws_unit, 3, 5, SUBHEADER_FILL, SUBHEADER_FONT)

unit_rows = [
    # (row, label, formula, unit, comment, section?)
    (5, "ДОХОД НА 1 АВТОМОБИЛЬ", None, None, None, True),
    (6, "Доход за 1 аренду (gross)", "=INPUTS!C12*INPUTS!C13", "€", "Цена/день × Дни аренды"),
    (7, "Количество аренд в месяц", "=INPUTS!C15", "шт.", "Из INPUTS"),
    (8, "Месячный доход (gross)", "=B6*B7", "€", "Доход за аренду × Кол-во аренд"),
    (9, "Комиссия агрегаторов", "=B8*INPUTS!C16", "€", "Месячный доход × % комиссии"),
    (10, "Месячный доход (net)", "=B8-B9", "€", "После вычета комиссии"),
    (11, "Годовой доход (net)", "=B10*12", "€", "Месячный × 12"),
    (13, "РАСХОДЫ НА 1 АВТОМОБИЛЬ (в месяц)", None, None, None, True),
    (14, "Амортизация (в месяц)", "=INPUTS!C7*(1-INPUTS!C9)/(INPUTS!C8*12)", "€", "Стоимость×(1-ост.%)/(Срок×12)"),
    (15, "Страховка", "=INPUTS!C19", "€", "Из INPUTS"),
    (16, "Ремонт и обслуживание", "=INPUTS!C20", "€", "Из INPUTS"),
    (17, "Парковка", "=INPUTS!C21", "€", "Из INPUTS"),
    (18, "Топливо / подготовка", "=INPUTS!C22", "€", "Из INPUTS"),
    (19, "ITV + налоги (в месяц)", "=INPUTS!C23/12", "€", "Годовая сумма / 12"),
    (20, "ИТОГО расходы на 1 авто / мес", "=SUM(B14:B19)", "€", "Сумма всех расходов"),
    (21, "ИТОГО расходы на 1 авто / год", "=B20*12", "€", "× 12 месяцев"),
    (23, "ПРИБЫЛЬ НА 1 АВТОМОБИЛЬ", None, None, None, True),
    (24, "Валовая прибыль (месяц)", "=B10-B20", "€", "Доход net - Расходы"),
    (25, "Валовая прибыль (год)", "=B24*12", "€", "× 12"),
    (26, "Маржинальность", "=B24/B10", "%", "Валовая прибыль / Доход"),
    (28, "ОКУПАЕМОСТЬ", None, None, None, True),
    (29, "Стоимость автомобиля", "=INPUTS!C7", "€", "Из INPUTS"),
    (30, "Чистая стоимость (за вычетом остат.)", "=INPUTS!C7*(1-INPUTS!C9)", "€", "Стоимость × (1 - ост.%)"),
    (31, "Срок окупаемости", "=IF(B24>0,B30/B24,\"N/A\")", "мес.", "Чистая стоимость / Валовая прибыль в мес."),
    (32, "ROI (годовой)", "=IF(B29>0,B25/B29,0)", "%", "Годовая прибыль / Стоимость авто"),
]

for row_num, label, formula, unit, comment, *section in unit_rows:
    is_section = section and section[0]
    ws_unit.cell(row=row_num, column=2, value=label)
    if formula:
        # formulas reference column C but we write to column C (col 3)
        actual_formula = formula.replace("=B", "=C")  # we'll use column C for values
        ws_unit.cell(row=row_num, column=3, value="").value = formula.replace("B", "C")
        if unit == "%":
            ws_unit.cell(row=row_num, column=3).number_format = PCT_FORMAT
        elif unit in ["€", "€/мес"]:
            ws_unit.cell(row=row_num, column=3).number_format = EUR_FORMAT
        elif unit == "мес.":
            ws_unit.cell(row=row_num, column=3).number_format = MONTHS_FORMAT
        else:
            ws_unit.cell(row=row_num, column=3).number_format = NUM_FORMAT
    if unit:
        ws_unit.cell(row=row_num, column=4, value=unit)
    if comment:
        ws_unit.cell(row=row_num, column=5, value=comment).font = Font(name="Calibri", size=9, italic=True, color="888888")

    if is_section:
        style_row(ws_unit, row_num, 5, SECTION_FILL, SECTION_FONT)
    else:
        style_row(ws_unit, row_num, 5, None, NORMAL_FONT)
        ws_unit.cell(row=row_num, column=3).fill = RESULT_FILL
        ws_unit.cell(row=row_num, column=3).font = RESULT_FONT

# Fix formulas to use column C properly
# Actually let me rewrite - formulas should reference column C in UNIT ECONOMICS
# Values are in column C (3)
for row_num, label, formula, unit, comment, *section in unit_rows:
    if formula:
        # Replace self-references (B -> C within this sheet)
        fixed = formula
        # Replace =B with =C for same-sheet refs
        import re
        # Fix internal refs: B6 -> C6, etc. but NOT INPUTS!C
        def fix_ref(m):
            return 'C' + m.group(1)
        fixed = re.sub(r'(?<!INPUTS!)B(\d+)', fix_ref, formula)
        # SUM(B14:B19) -> SUM(C14:C19)
        fixed = fixed.replace('SUM(B', 'SUM(C').replace(':B', ':C')
        ws_unit.cell(row=row_num, column=3).value = fixed


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: FLEET PERFORMANCE
# ═══════════════════════════════════════════════════════════════════════════
ws_fleet = wb.create_sheet("FLEET PERFORMANCE")
ws_fleet.sheet_properties.tabColor = "00B050"
set_col_widths(ws_fleet, [5, 45, 20, 15, 35])

ws_fleet.merge_cells('B1:D1')
ws_fleet['B1'] = "FLEET PERFORMANCE — Показатели автопарка"
ws_fleet['B1'].font = TITLE_FONT

for col, val in [(2, "Показатель"), (3, "Значение"), (4, "Ед."), (5, "Формула")]:
    ws_fleet.cell(row=3, column=col, value=val)
style_header(ws_fleet, 3, 5, PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"), HEADER_FONT)

fleet_data = [
    (5, "ОБЩИЕ ПОКАЗАТЕЛИ ФЛОТА", None, None, None, True),
    (6, "Количество автомобилей", "=INPUTS!C6", "шт.", "Из INPUTS"),
    (7, "Загрузка автопарка", "=INPUTS!C14", "%", "Из INPUTS"),
    (8, "Дней аренды на 1 авто в месяц", "=30*INPUTS!C14", "дней", "30 × Загрузка"),
    (9, "Дней аренды всего флота в месяц", "=C8*C6", "дней", "Дней на авто × Кол-во авто"),
    (11, "АРЕНДЫ", None, None, None, True),
    (12, "Аренд на 1 авто в месяц", "=INPUTS!C15", "шт.", "Из INPUTS"),
    (13, "Аренд на 1 авто в день", "=C12/30", "шт.", "В месяц / 30"),
    (14, "Аренд всего флота в день", "=C13*C6", "шт.", "На 1 авто в день × Флот"),
    (15, "Аренд всего флота в месяц", "=C12*C6", "шт.", "На 1 авто в мес × Флот"),
    (16, "Аренд всего флота в год", "=C15*12", "шт.", "В месяц × 12"),
    (18, "ДОХОД ФЛОТА", None, None, None, True),
    (19, "Доход за 1 аренду (gross)", "=INPUTS!C12*INPUTS!C13", "€", "Цена/день × Дни"),
    (20, "Месячный доход флота (gross)", "=C19*C15", "€", "Доход за аренду × Аренды в мес"),
    (21, "Комиссия агрегаторов (мес)", "=C20*INPUTS!C16", "€", "Gross × % комиссии"),
    (22, "Месячный доход флота (net)", "=C20-C21", "€", "Gross - Комиссия"),
    (23, "Годовой доход флота (net)", "=C22*12", "€", "× 12"),
    (24, "Средний доход на 1 авто (мес, net)", "=C22/C6", "€", "Доход net / Кол-во авто"),
    (25, "Средний доход на 1 авто (год, net)", "=C24*12", "€", "× 12"),
    (27, "РАСХОДЫ ФЛОТА", None, None, None, True),
    (28, "Расходы на весь флот (мес)", "='UNIT ECONOMICS'!C20*C6", "€", "Расходы на 1 авто × Флот"),
    (29, "Расходы на весь флот (год)", "=C28*12", "€", "× 12"),
    (31, "ВАЛОВАЯ ПРИБЫЛЬ ФЛОТА", None, None, None, True),
    (32, "Валовая прибыль флота (мес)", "=C22-C28", "€", "Доход net - Расходы"),
    (33, "Валовая прибыль флота (год)", "=C32*12", "€", "× 12"),
    (34, "Маржинальность флота", "=IF(C22>0,C32/C22,0)", "%", "Валовая прибыль / Доход"),
]

for row_num, label, formula, unit, comment, *section in fleet_data:
    is_section = section and section[0]
    ws_fleet.cell(row=row_num, column=2, value=label)
    if formula:
        ws_fleet.cell(row=row_num, column=3, value=formula)
        if unit == "%":
            ws_fleet.cell(row=row_num, column=3).number_format = PCT_FORMAT
        elif unit == "€":
            ws_fleet.cell(row=row_num, column=3).number_format = EUR_FORMAT
        else:
            ws_fleet.cell(row=row_num, column=3).number_format = NUM_FORMAT
    if unit:
        ws_fleet.cell(row=row_num, column=4, value=unit)
    if comment:
        ws_fleet.cell(row=row_num, column=5, value=comment).font = Font(name="Calibri", size=9, italic=True, color="888888")
    if is_section:
        style_row(ws_fleet, row_num, 5, SECTION_FILL, SECTION_FONT)
    else:
        style_row(ws_fleet, row_num, 5, None, NORMAL_FONT)
        if formula:
            ws_fleet.cell(row=row_num, column=3).fill = RESULT_FILL
            ws_fleet.cell(row=row_num, column=3).font = RESULT_FONT


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: REVENUE FORECAST
# ═══════════════════════════════════════════════════════════════════════════
ws_rev = wb.create_sheet("REVENUE FORECAST")
ws_rev.sheet_properties.tabColor = "FFC000"
set_col_widths(ws_rev, [5, 25] + [16]*13 + [18])

ws_rev.merge_cells('B1:N1')
ws_rev['B1'] = "REVENUE FORECAST — Прогноз доходов на 12 месяцев"
ws_rev['B1'].font = TITLE_FONT

months_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Headers
ws_rev.cell(row=3, column=2, value="Показатель")
for i, m in enumerate(months_es):
    ws_rev.cell(row=3, column=3+i, value=m)
ws_rev.cell(row=3, column=15, value="ИТОГО ГОД")
style_header(ws_rev, 3, 15, PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
             Font(name="Calibri", bold=True, color="000000", size=10))

# Seasonality: high season = Jun-Oct (months 6-10), rest is low
# Row 5: Сезонный коэффициент
ws_rev.cell(row=4, column=2, value="СЕЗОННОСТЬ И ЗАГРУЗКА")
style_row(ws_rev, 4, 15, SECTION_FILL, SECTION_FONT)

ws_rev.cell(row=5, column=2, value="Сезонный коэффициент")
for i in range(12):
    col = 3 + i
    month_num = i + 1
    # High season: June(6), July(7), August(8), September(9), October(10)
    formula = f'=IF(OR({month_num}=6,{month_num}=7,{month_num}=8,{month_num}=9,{month_num}=10),INPUTS!C38,INPUTS!C39)'
    ws_rev.cell(row=5, column=col, value=formula)
    ws_rev.cell(row=5, column=col).number_format = '0.00'

ws_rev.cell(row=6, column=2, value="Загрузка (с учётом сезона)")
for i in range(12):
    col = 3 + i
    ws_rev.cell(row=6, column=col, value=f'=INPUTS!C14*{get_column_letter(col)}5')
    ws_rev.cell(row=6, column=col).number_format = PCT_FORMAT

ws_rev.cell(row=7, column=2, value="Кол-во авто")
for i in range(12):
    col = 3 + i
    # Linear growth throughout year
    ws_rev.cell(row=7, column=col, value=f'=ROUND(INPUTS!C6*(1+INPUTS!C36*{i}/12),0)')
    ws_rev.cell(row=7, column=col).number_format = NUM_FORMAT

# Revenue rows
ws_rev.cell(row=9, column=2, value="ДОХОДЫ")
style_row(ws_rev, 9, 15, SECTION_FILL, SECTION_FONT)

ws_rev.cell(row=10, column=2, value="Дней аренды (флот)")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=10, column=col, value=f'=30*{cl}6*{cl}7')
    ws_rev.cell(row=10, column=col).number_format = NUM_FORMAT

ws_rev.cell(row=11, column=2, value="Выручка (gross)")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=11, column=col, value=f'={cl}10*INPUTS!C12')
    ws_rev.cell(row=11, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=12, column=2, value="Комиссия агрегаторов")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=12, column=col, value=f'={cl}11*INPUTS!C16')
    ws_rev.cell(row=12, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=13, column=2, value="Выручка (net)")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=13, column=col, value=f'={cl}11-{cl}12')
    ws_rev.cell(row=13, column=col).number_format = EUR_FORMAT_0
    ws_rev.cell(row=13, column=col).font = BOLD_FONT

# Year totals
for r in [10, 11, 12, 13]:
    ws_rev.cell(row=r, column=15, value=f'=SUM(C{r}:N{r})')
    ws_rev.cell(row=r, column=15).number_format = EUR_FORMAT_0 if r > 10 else NUM_FORMAT
    ws_rev.cell(row=r, column=15).font = BOLD_FONT

# Costs rows
ws_rev.cell(row=15, column=2, value="РАСХОДЫ")
style_row(ws_rev, 15, 15, SECTION_FILL, SECTION_FONT)

ws_rev.cell(row=16, column=2, value="Расходы на авто (флот)")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=16, column=col, value=f"='UNIT ECONOMICS'!C20*{cl}7")
    ws_rev.cell(row=16, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=17, column=2, value="ФОТ (с налогами)")
for i in range(12):
    col = 3 + i
    ws_rev.cell(row=17, column=col, value="=INPUTS!C26*INPUTS!C27*(1+INPUTS!C28)")
    ws_rev.cell(row=17, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=18, column=2, value="Аренда офиса")
for i in range(12):
    col = 3 + i
    ws_rev.cell(row=18, column=col, value="=INPUTS!C29")
    ws_rev.cell(row=18, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=19, column=2, value="Маркетинг")
for i in range(12):
    col = 3 + i
    ws_rev.cell(row=19, column=col, value="=INPUTS!C30")
    ws_rev.cell(row=19, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=20, column=2, value="Софт + Связь + Прочие")
for i in range(12):
    col = 3 + i
    ws_rev.cell(row=20, column=col, value="=INPUTS!C31+INPUTS!C32+INPUTS!C33")
    ws_rev.cell(row=20, column=col).number_format = EUR_FORMAT_0

ws_rev.cell(row=21, column=2, value="ИТОГО расходы")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=21, column=col, value=f'=SUM({cl}16:{cl}20)')
    ws_rev.cell(row=21, column=col).number_format = EUR_FORMAT_0
    ws_rev.cell(row=21, column=col).font = BOLD_FONT

for r in [16, 17, 18, 19, 20, 21]:
    ws_rev.cell(row=r, column=15, value=f'=SUM(C{r}:N{r})')
    ws_rev.cell(row=r, column=15).number_format = EUR_FORMAT_0
    ws_rev.cell(row=r, column=15).font = BOLD_FONT

# Profit
ws_rev.cell(row=23, column=2, value="ПРИБЫЛЬ")
style_row(ws_rev, 23, 15, SECTION_FILL, SECTION_FONT)

ws_rev.cell(row=24, column=2, value="EBITDA")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=24, column=col, value=f'={cl}13-{cl}21')
    ws_rev.cell(row=24, column=col).number_format = EUR_FORMAT_0
    ws_rev.cell(row=24, column=col).font = BOLD_FONT

ws_rev.cell(row=24, column=15, value=f'=SUM(C24:N24)')
ws_rev.cell(row=24, column=15).number_format = EUR_FORMAT_0
ws_rev.cell(row=24, column=15).font = BOLD_FONT

ws_rev.cell(row=25, column=2, value="Маржа EBITDA")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=25, column=col, value=f'=IF({cl}13>0,{cl}24/{cl}13,0)')
    ws_rev.cell(row=25, column=col).number_format = PCT_FORMAT

ws_rev.cell(row=25, column=15, value='=IF(O13>0,O24/O13,0)')
ws_rev.cell(row=25, column=15).number_format = PCT_FORMAT

# Quarterly summary
ws_rev.cell(row=28, column=2, value="КВАРТАЛЬНАЯ СВОДКА")
style_row(ws_rev, 28, 7, SECTION_FILL, SECTION_FONT)
for col, val in [(2, ""), (3, "Q1"), (4, "Q2"), (5, "Q3"), (6, "Q4"), (7, "ГОД")]:
    ws_rev.cell(row=29, column=col, value=val)
style_header(ws_rev, 29, 7, SUBHEADER_FILL, SUBHEADER_FONT)

q_rows = [
    (30, "Выручка net", 13),
    (31, "Расходы", 21),
    (32, "EBITDA", 24),
]
for r, label, src_row in q_rows:
    ws_rev.cell(row=r, column=2, value=label)
    ws_rev.cell(row=r, column=3, value=f'=SUM(C{src_row}:E{src_row})')  # Q1
    ws_rev.cell(row=r, column=4, value=f'=SUM(F{src_row}:H{src_row})')  # Q2
    ws_rev.cell(row=r, column=5, value=f'=SUM(I{src_row}:K{src_row})')  # Q3
    ws_rev.cell(row=r, column=6, value=f'=SUM(L{src_row}:N{src_row})')  # Q4
    ws_rev.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')             # Year
    for c in range(3, 8):
        ws_rev.cell(row=r, column=c).number_format = EUR_FORMAT_0
        ws_rev.cell(row=r, column=c).font = BOLD_FONT
        ws_rev.cell(row=r, column=c).border = thin_border
    ws_rev.cell(row=r, column=2).border = thin_border


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: P&L
# ═══════════════════════════════════════════════════════════════════════════
ws_pl = wb.create_sheet("P&L")
ws_pl.sheet_properties.tabColor = "FF0000"
set_col_widths(ws_pl, [5, 40, 20, 15, 35])

ws_pl.merge_cells('B1:D1')
ws_pl['B1'] = "P&L — Отчёт о прибылях и убытках (годовой)"
ws_pl['B1'].font = TITLE_FONT

for col, val in [(2, "Статья"), (3, "Сумма (год)"), (4, "% от выручки"), (5, "Комментарий")]:
    ws_pl.cell(row=3, column=col, value=val)
style_header(ws_pl, 3, 5, PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"), HEADER_FONT)

pl_items = [
    (5, "ВЫРУЧКА", None, None, None, True),
    (6, "Выручка (gross)", "='FLEET PERFORMANCE'!C23/(1-INPUTS!C16)", None, "Годовой доход флота до комиссии"),
    (7, "Комиссия агрегаторов", "=C6*INPUTS!C16", None, "% комиссии × Gross"),
    (8, "Выручка (net)", "=C6-C7", None, "Gross - Комиссии"),
    (10, "СЕБЕСТОИМОСТЬ (COGS)", None, None, None, True),
    (11, "Амортизация автопарка", "=INPUTS!C6*INPUTS!C7*(1-INPUTS!C9)/INPUTS!C8", None, "Амортизация всего флота за год"),
    (12, "Страховка", "=INPUTS!C19*INPUTS!C6*12", None, "На авто × Флот × 12"),
    (13, "Ремонт и обслуживание", "=INPUTS!C20*INPUTS!C6*12", None, ""),
    (14, "Парковка", "=INPUTS!C21*INPUTS!C6*12", None, ""),
    (15, "Топливо / подготовка", "=INPUTS!C22*INPUTS!C6*12", None, ""),
    (16, "ITV + налоги", "=INPUTS!C23*INPUTS!C6", None, ""),
    (17, "ИТОГО себестоимость", "=SUM(C11:C16)", None, ""),
    (19, "ВАЛОВАЯ ПРИБЫЛЬ", "=C8-C17", None, "", True),
    (20, "Маржинальность валовой прибыли", None, "=IF(C8>0,C19/C8,0)", "Валовая прибыль / Выручка net", False),
    (22, "ОПЕРАЦИОННЫЕ РАСХОДЫ (OPEX)", None, None, None, True),
    (23, "ФОТ (с налогами на зарплату)", "=INPUTS!C26*INPUTS!C27*(1+INPUTS!C28)*12", None, "Сотрудники × ЗП × (1+налог) × 12"),
    (24, "Аренда офиса / точки", "=INPUTS!C29*12", None, "× 12 месяцев"),
    (25, "Маркетинг и реклама", "=INPUTS!C30*12", None, "× 12 месяцев"),
    (26, "Софт и IT", "=INPUTS!C31*12", None, ""),
    (27, "Связь и интернет", "=INPUTS!C32*12", None, ""),
    (28, "Прочие расходы", "=INPUTS!C33*12", None, ""),
    (29, "ИТОГО OPEX", "=SUM(C23:C28)", None, ""),
    (31, "EBITDA", "=C19-C29", None, "", True),
    (32, "Маржа EBITDA", None, "=IF(C8>0,C31/C8,0)", "", False),
    (34, "Амортизация (уже учтена в COGS)", "=C11", None, "Для справки", False),
    (35, "EBIT", "=C31", None, "= EBITDA (аморт. уже в COGS)", False),
    (37, "НАЛОГ НА ПРИБЫЛЬ", "=IF(C35>0,C35*INPUTS!C42,0)", None, "Ставка из INPUTS", False),
    (39, "ЧИСТАЯ ПРИБЫЛЬ", "=C35-C37", None, "", True),
    (40, "Рентабельность чистой прибыли", None, "=IF(C8>0,C39/C8,0)", "", False),
]

for row_num, label, formula_c, formula_d, comment, *section in pl_items:
    is_section = section and section[0]
    ws_pl.cell(row=row_num, column=2, value=label).font = BOLD_FONT if is_section else NORMAL_FONT
    if formula_c:
        ws_pl.cell(row=row_num, column=3, value=formula_c)
        ws_pl.cell(row=row_num, column=3).number_format = EUR_FORMAT
    if formula_d:
        ws_pl.cell(row=row_num, column=4, value=formula_d)
        ws_pl.cell(row=row_num, column=4).number_format = PCT_FORMAT
    elif formula_c:
        ws_pl.cell(row=row_num, column=4, value=f'=IF($C$8>0,C{row_num}/$C$8,0)')
        ws_pl.cell(row=row_num, column=4).number_format = PCT_FORMAT
    if comment:
        ws_pl.cell(row=row_num, column=5, value=comment).font = Font(name="Calibri", size=9, italic=True, color="888888")

    if is_section:
        style_row(ws_pl, row_num, 5, SECTION_FILL, SECTION_FONT)
    else:
        for c in range(2, 6):
            ws_pl.cell(row=row_num, column=c).border = thin_border

# Highlight key rows
for r in [8, 19, 31, 39]:
    ws_pl.cell(row=r, column=3).fill = RESULT_FILL
    ws_pl.cell(row=r, column=3).font = RESULT_FONT


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: CASH FLOW
# ═══════════════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("CASH FLOW")
ws_cf.sheet_properties.tabColor = "7030A0"
set_col_widths(ws_cf, [5, 40] + [16]*13)

ws_cf.merge_cells('B1:N1')
ws_cf['B1'] = "CASH FLOW — Движение денежных средств (12 месяцев)"
ws_cf['B1'].font = TITLE_FONT

ws_cf.cell(row=3, column=2, value="Статья")
for i, m in enumerate(months_es):
    ws_cf.cell(row=3, column=3+i, value=m)
ws_cf.cell(row=3, column=15, value="ИТОГО")
style_header(ws_cf, 3, 15, PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"), HEADER_FONT)

# Operating Cash In
ws_cf.cell(row=5, column=2, value="ОПЕРАЦИОННЫЕ ПОСТУПЛЕНИЯ")
style_row(ws_cf, 5, 15, SECTION_FILL, SECTION_FONT)

ws_cf.cell(row=6, column=2, value="Выручка от аренды (net)")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=6, column=col, value=f"='REVENUE FORECAST'!{cl}13")
    ws_cf.cell(row=6, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=7, column=2, value="Залоги (депозиты) — приход")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    # Estimate: average deposit 300€ × number of rentals per month
    ws_cf.cell(row=7, column=col, value=f"=300*'REVENUE FORECAST'!{cl}7*INPUTS!C15")
    ws_cf.cell(row=7, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=8, column=2, value="Возврат залогов")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=8, column=col, value=f"=-{cl}7*0.95")  # 95% returned
    ws_cf.cell(row=8, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=9, column=2, value="ИТОГО поступления")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=9, column=col, value=f"={cl}6+{cl}7+{cl}8")
    ws_cf.cell(row=9, column=col).number_format = EUR_FORMAT_0
    ws_cf.cell(row=9, column=col).font = BOLD_FONT

# Operating Cash Out
ws_cf.cell(row=11, column=2, value="ОПЕРАЦИОННЫЕ РАСХОДЫ")
style_row(ws_cf, 11, 15, SECTION_FILL, SECTION_FONT)

cf_expense_rows = [
    (12, "Расходы на автопарк", "='REVENUE FORECAST'!{cl}16"),
    (13, "ФОТ", "='REVENUE FORECAST'!{cl}17"),
    (14, "Аренда офиса", "='REVENUE FORECAST'!{cl}18"),
    (15, "Маркетинг", "='REVENUE FORECAST'!{cl}19"),
    (16, "Софт + Связь + Прочие", "='REVENUE FORECAST'!{cl}20"),
]

for r, label, formula_tmpl in cf_expense_rows:
    ws_cf.cell(row=r, column=2, value=label)
    for i in range(12):
        col = 3 + i
        cl = get_column_letter(col)
        ws_cf.cell(row=r, column=col, value=f"=-{formula_tmpl.format(cl=cl)[1:]}")
        ws_cf.cell(row=r, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=17, column=2, value="Налог на прибыль (квартальный)")
for i in range(12):
    col = 3 + i
    month = i + 1
    # Tax paid quarterly (months 3, 6, 9, 12)
    if month in [3, 6, 9, 12]:
        ws_cf.cell(row=17, column=col, value=f"=-'P&L'!C37/4")
    else:
        ws_cf.cell(row=17, column=col, value=0)
    ws_cf.cell(row=17, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=18, column=2, value="ИТОГО расходы")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=18, column=col, value=f"=SUM({cl}12:{cl}17)")
    ws_cf.cell(row=18, column=col).number_format = EUR_FORMAT_0
    ws_cf.cell(row=18, column=col).font = BOLD_FONT

# Investment
ws_cf.cell(row=20, column=2, value="ИНВЕСТИЦИИ")
style_row(ws_cf, 20, 15, SECTION_FILL, SECTION_FONT)

ws_cf.cell(row=21, column=2, value="Покупка новых автомобилей")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    # New cars added each month (from growth)
    if i == 0:
        ws_cf.cell(row=21, column=col, value=0)
    else:
        prev_cl = get_column_letter(col - 1)
        ws_cf.cell(row=21, column=col,
                   value=f"=-MAX(0,'REVENUE FORECAST'!{cl}7-'REVENUE FORECAST'!{prev_cl}7)*INPUTS!C7")
    ws_cf.cell(row=21, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=22, column=2, value="Продажа старых авто (выручка)")
for i in range(12):
    col = 3 + i
    ws_cf.cell(row=22, column=col, value=0)  # Manual input if needed
    ws_cf.cell(row=22, column=col).number_format = EUR_FORMAT_0
    ws_cf.cell(row=22, column=col).fill = INPUT_FILL

# Net Cash Flow
ws_cf.cell(row=24, column=2, value="ДЕНЕЖНЫЙ ПОТОК")
style_row(ws_cf, 24, 15, SECTION_FILL, SECTION_FONT)

ws_cf.cell(row=25, column=2, value="Операционный денежный поток")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=25, column=col, value=f"={cl}9+{cl}18")
    ws_cf.cell(row=25, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=26, column=2, value="Инвестиционный денежный поток")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=26, column=col, value=f"={cl}21+{cl}22")
    ws_cf.cell(row=26, column=col).number_format = EUR_FORMAT_0

ws_cf.cell(row=27, column=2, value="ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_cf.cell(row=27, column=col, value=f"={cl}25+{cl}26")
    ws_cf.cell(row=27, column=col).number_format = EUR_FORMAT_0
    ws_cf.cell(row=27, column=col).font = BOLD_FONT

ws_cf.cell(row=29, column=2, value="Накопительный денежный поток")
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    if i == 0:
        ws_cf.cell(row=29, column=col, value=f"={cl}27")
    else:
        prev_cl = get_column_letter(col - 1)
        ws_cf.cell(row=29, column=col, value=f"={prev_cl}29+{cl}27")
    ws_cf.cell(row=29, column=col).number_format = EUR_FORMAT_0
    ws_cf.cell(row=29, column=col).font = BOLD_FONT

# Totals column
for r in [6, 7, 8, 9, 12, 13, 14, 15, 16, 17, 18, 21, 22, 25, 26, 27]:
    ws_cf.cell(row=r, column=15, value=f'=SUM(C{r}:N{r})')
    ws_cf.cell(row=r, column=15).number_format = EUR_FORMAT_0
    ws_cf.cell(row=r, column=15).font = BOLD_FONT


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("DASHBOARD")
ws_dash.sheet_properties.tabColor = "00B0F0"
set_col_widths(ws_dash, [3, 25, 20, 5, 25, 20, 5, 25, 20])

ws_dash.merge_cells('B1:H1')
ws_dash['B1'] = "DASHBOARD — Панель управления VoyageRent"
ws_dash['B1'].font = TITLE_FONT

# KPI Cards - Row 1
kpi_data = [
    # (col_label, col_value, label, formula, fmt)
    (2, 3, "Автопарк", "=INPUTS!C6", NUM_FORMAT),
    (5, 6, "Загрузка", "=INPUTS!C14", PCT_FORMAT),
    (8, 9, "Цена аренды / день", "=INPUTS!C12", EUR_FORMAT),
]

for i, (cl, cv, label, formula, fmt) in enumerate(kpi_data):
    r = 3
    ws_dash.cell(row=r, column=cl, value=label)
    ws_dash.cell(row=r, column=cl).font = SUBHEADER_FONT
    ws_dash.cell(row=r, column=cl).fill = SUBHEADER_FILL
    ws_dash.cell(row=r, column=cv, value="")
    ws_dash.cell(row=r, column=cv).fill = SUBHEADER_FILL
    ws_dash.cell(row=r+1, column=cl, value="")
    ws_dash.cell(row=r+1, column=cl).fill = KPI_FILL
    ws_dash.cell(row=r+1, column=cv, value=formula)
    ws_dash.cell(row=r+1, column=cv).font = Font(name="Calibri", bold=True, size=18, color="1B3A5C")
    ws_dash.cell(row=r+1, column=cv).number_format = fmt
    ws_dash.cell(row=r+1, column=cv).fill = KPI_FILL

# KPI Cards - Row 2
kpi_data2 = [
    (2, 3, "Выручка (год, net)", "='P&L'!C8", EUR_FORMAT_0),
    (5, 6, "EBITDA (год)", "='P&L'!C31", EUR_FORMAT_0),
    (8, 9, "Чистая прибыль (год)", "='P&L'!C39", EUR_FORMAT_0),
]

for cl, cv, label, formula, fmt in kpi_data2:
    r = 6
    ws_dash.cell(row=r, column=cl, value=label)
    ws_dash.cell(row=r, column=cl).font = SUBHEADER_FONT
    ws_dash.cell(row=r, column=cl).fill = SUBHEADER_FILL
    ws_dash.cell(row=r, column=cv).fill = SUBHEADER_FILL
    ws_dash.cell(row=r+1, column=cl).fill = KPI_FILL
    ws_dash.cell(row=r+1, column=cv, value=formula)
    ws_dash.cell(row=r+1, column=cv).font = Font(name="Calibri", bold=True, size=18, color="1B3A5C")
    ws_dash.cell(row=r+1, column=cv).number_format = fmt
    ws_dash.cell(row=r+1, column=cv).fill = KPI_FILL

# KPI Cards - Row 3
kpi_data3 = [
    (2, 3, "Доход на 1 авто (мес)", "='UNIT ECONOMICS'!C10", EUR_FORMAT),
    (5, 6, "Прибыль на 1 авто (мес)", "='UNIT ECONOMICS'!C24", EUR_FORMAT),
    (8, 9, "Срок окупаемости авто", "='UNIT ECONOMICS'!C31", '0.0 "мес."'),
]

for cl, cv, label, formula, fmt in kpi_data3:
    r = 9
    ws_dash.cell(row=r, column=cl, value=label)
    ws_dash.cell(row=r, column=cl).font = SUBHEADER_FONT
    ws_dash.cell(row=r, column=cl).fill = SUBHEADER_FILL
    ws_dash.cell(row=r, column=cv).fill = SUBHEADER_FILL
    ws_dash.cell(row=r+1, column=cl).fill = KPI_FILL
    ws_dash.cell(row=r+1, column=cv, value=formula)
    ws_dash.cell(row=r+1, column=cv).font = Font(name="Calibri", bold=True, size=18, color="1B3A5C")
    ws_dash.cell(row=r+1, column=cv).number_format = fmt
    ws_dash.cell(row=r+1, column=cv).fill = KPI_FILL

# KPI Cards - Row 4
kpi_data4 = [
    (2, 3, "ROI на автомобиль", "='UNIT ECONOMICS'!C32", PCT_FORMAT),
    (5, 6, "Маржа EBITDA", "='P&L'!D32", PCT_FORMAT),
    (8, 9, "Аренд в день (флот)", "='FLEET PERFORMANCE'!C14", '0.0'),
]

for cl, cv, label, formula, fmt in kpi_data4:
    r = 12
    ws_dash.cell(row=r, column=cl, value=label)
    ws_dash.cell(row=r, column=cl).font = SUBHEADER_FONT
    ws_dash.cell(row=r, column=cl).fill = SUBHEADER_FILL
    ws_dash.cell(row=r, column=cv).fill = SUBHEADER_FILL
    ws_dash.cell(row=r+1, column=cl).fill = KPI_FILL
    ws_dash.cell(row=r+1, column=cv, value=formula)
    ws_dash.cell(row=r+1, column=cv).font = Font(name="Calibri", bold=True, size=18, color="1B3A5C")
    ws_dash.cell(row=r+1, column=cv).number_format = fmt
    ws_dash.cell(row=r+1, column=cv).fill = KPI_FILL

# Charts
# Chart 1: Monthly Revenue (Bar)
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Месячная выручка (net)"
chart1.y_axis.title = "EUR"
chart1.x_axis.title = "Месяц"
chart1.style = 10
data1 = Reference(ws_rev, min_col=3, min_row=13, max_col=14, max_row=13)
cats1 = Reference(ws_rev, min_col=3, min_row=3, max_col=14, max_row=3)
chart1.add_data(data1, from_rows=True, titles_from_data=False)
chart1.set_categories(cats1)
chart1.shape = 4
chart1.width = 28
chart1.height = 14
ws_dash.add_chart(chart1, "B16")

# Chart 2: Monthly EBITDA (Line)
chart2 = LineChart()
chart2.title = "EBITDA по месяцам"
chart2.y_axis.title = "EUR"
chart2.style = 10
data2 = Reference(ws_rev, min_col=3, min_row=24, max_col=14, max_row=24)
chart2.add_data(data2, from_rows=True, titles_from_data=False)
chart2.set_categories(cats1)
chart2.width = 28
chart2.height = 14
ws_dash.add_chart(chart2, "B32")

# Chart 3: Cost Structure (Pie) from P&L
chart3 = PieChart()
chart3.title = "Структура расходов"
chart3.style = 10
# We'll create a small helper range for the pie chart
ws_dash.cell(row=50, column=2, value="COGS")
ws_dash.cell(row=50, column=3, value="='P&L'!C17")
ws_dash.cell(row=51, column=2, value="OPEX")
ws_dash.cell(row=51, column=3, value="='P&L'!C29")
ws_dash.cell(row=52, column=2, value="Налоги")
ws_dash.cell(row=52, column=3, value="='P&L'!C37")
labels = Reference(ws_dash, min_col=2, min_row=50, max_row=52)
values = Reference(ws_dash, min_col=3, min_row=50, max_row=52)
chart3.add_data(values, titles_from_data=False)
chart3.set_categories(labels)
chart3.width = 18
chart3.height = 14
ws_dash.add_chart(chart3, "B48")


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8: SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════
ws_sc = wb.create_sheet("SCENARIOS")
ws_sc.sheet_properties.tabColor = "ED7D31"
set_col_widths(ws_sc, [5, 35, 20, 20, 20, 5, 35])

ws_sc.merge_cells('B1:E1')
ws_sc['B1'] = "SCENARIOS — Сценарный анализ"
ws_sc['B1'].font = TITLE_FONT

ws_sc.merge_cells('B2:E2')
ws_sc['B2'] = "Измените параметры сценариев и сравните результаты"
ws_sc['B2'].font = Font(name="Calibri", italic=True, size=10, color="666666")

# Scenario Parameters
ws_sc.cell(row=4, column=2, value="ПАРАМЕТРЫ СЦЕНАРИЕВ")
style_row(ws_sc, 4, 5, SECTION_FILL, SECTION_FONT)

for col, val in [(2, "Параметр"), (3, "Pessimistic"), (4, "Base"), (5, "Optimistic")]:
    ws_sc.cell(row=5, column=col, value=val)
style_header(ws_sc, 5, 5, PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
             Font(name="Calibri", bold=True, color="FFFFFF", size=11))

scenario_params = [
    (6, "Количество автомобилей", 80, "=INPUTS!C6", 130),
    (7, "Цена аренды (€/день)", 35, "=INPUTS!C12", 55),
    (8, "Загрузка (%)", 0.50, "=INPUTS!C14", 0.80),
    (9, "Аренд в месяц (на 1 авто)", 3, "=INPUTS!C15", 5),
    (10, "Комиссия агрегаторов (%)", 0.20, "=INPUTS!C16", 0.10),
]

for r, label, pess, base, opt in scenario_params:
    ws_sc.cell(row=r, column=2, value=label).font = NORMAL_FONT
    ws_sc.cell(row=r, column=2).border = thin_border
    for col, val in [(3, pess), (4, base), (5, opt)]:
        c = ws_sc.cell(row=r, column=col, value=val)
        c.fill = INPUT_FILL
        c.font = INPUT_FONT
        c.border = thin_border
        if isinstance(val, float) and val < 1:
            c.number_format = PCT_FORMAT
        elif isinstance(val, str) and 'C14' in val:
            c.number_format = PCT_FORMAT
        elif isinstance(val, str) and 'C16' in val:
            c.number_format = PCT_FORMAT

# Results for each scenario
ws_sc.cell(row=12, column=2, value="РЕЗУЛЬТАТЫ СЦЕНАРИЕВ")
style_row(ws_sc, 12, 5, SECTION_FILL, SECTION_FONT)

for col, val in [(2, "Показатель"), (3, "Pessimistic"), (4, "Base"), (5, "Optimistic")]:
    ws_sc.cell(row=13, column=col, value=val)
style_header(ws_sc, 13, 5, SUBHEADER_FILL, SUBHEADER_FONT)

# For each scenario column (C=3, D=4, E=5):
for sc_col_letter, sc_col in [('C', 3), ('D', 4), ('E', 5)]:
    # Revenue per rental
    ws_sc.cell(row=14, column=2, value="Доход за 1 аренду")
    ws_sc.cell(row=14, column=sc_col, value=f"={sc_col_letter}7*INPUTS!C13")
    ws_sc.cell(row=14, column=sc_col).number_format = EUR_FORMAT

    # Monthly revenue per car (gross)
    ws_sc.cell(row=15, column=2, value="Мес. доход на авто (gross)")
    ws_sc.cell(row=15, column=sc_col, value=f"={sc_col_letter}14*{sc_col_letter}9")
    ws_sc.cell(row=15, column=sc_col).number_format = EUR_FORMAT

    # Monthly revenue per car (net)
    ws_sc.cell(row=16, column=2, value="Мес. доход на авто (net)")
    ws_sc.cell(row=16, column=sc_col, value=f"={sc_col_letter}15*(1-{sc_col_letter}10)")
    ws_sc.cell(row=16, column=sc_col).number_format = EUR_FORMAT

    # Monthly costs per car
    ws_sc.cell(row=17, column=2, value="Мес. расходы на авто")
    ws_sc.cell(row=17, column=sc_col, value=f"='UNIT ECONOMICS'!C20")
    ws_sc.cell(row=17, column=sc_col).number_format = EUR_FORMAT

    # Gross profit per car
    ws_sc.cell(row=18, column=2, value="Валовая прибыль на авто (мес)")
    ws_sc.cell(row=18, column=sc_col, value=f"={sc_col_letter}16-{sc_col_letter}17")
    ws_sc.cell(row=18, column=sc_col).number_format = EUR_FORMAT

    # Total fleet revenue (year, net)
    ws_sc.cell(row=19, column=2, value="Выручка флота (год, net)")
    ws_sc.cell(row=19, column=sc_col, value=f"={sc_col_letter}16*{sc_col_letter}6*12")
    ws_sc.cell(row=19, column=sc_col).number_format = EUR_FORMAT_0

    # Total fleet costs (year)
    ws_sc.cell(row=20, column=2, value="Расходы на флот (год)")
    ws_sc.cell(row=20, column=sc_col, value=f"={sc_col_letter}17*{sc_col_letter}6*12")
    ws_sc.cell(row=20, column=sc_col).number_format = EUR_FORMAT_0

    # OPEX (year)
    ws_sc.cell(row=21, column=2, value="OPEX (год)")
    ws_sc.cell(row=21, column=sc_col, value=f"='P&L'!C29")
    ws_sc.cell(row=21, column=sc_col).number_format = EUR_FORMAT_0

    # EBITDA
    ws_sc.cell(row=22, column=2, value="EBITDA (год)")
    ws_sc.cell(row=22, column=sc_col, value=f"={sc_col_letter}19-{sc_col_letter}20-{sc_col_letter}21")
    ws_sc.cell(row=22, column=sc_col).number_format = EUR_FORMAT_0

    # Net Profit
    ws_sc.cell(row=23, column=2, value="Чистая прибыль (год)")
    ws_sc.cell(row=23, column=sc_col, value=f"={sc_col_letter}22*(1-INPUTS!C42)")
    ws_sc.cell(row=23, column=sc_col).number_format = EUR_FORMAT_0

    # EBITDA margin
    ws_sc.cell(row=24, column=2, value="Маржа EBITDA")
    ws_sc.cell(row=24, column=sc_col, value=f"=IF({sc_col_letter}19>0,{sc_col_letter}22/{sc_col_letter}19,0)")
    ws_sc.cell(row=24, column=sc_col).number_format = PCT_FORMAT

    # Payback
    ws_sc.cell(row=25, column=2, value="Срок окупаемости авто (мес)")
    ws_sc.cell(row=25, column=sc_col,
               value=f"=IF({sc_col_letter}18>0,INPUTS!C7*(1-INPUTS!C9)/{sc_col_letter}18,\"N/A\")")
    ws_sc.cell(row=25, column=sc_col).number_format = MONTHS_FORMAT

    # ROI
    ws_sc.cell(row=26, column=2, value="ROI на автомобиль (год)")
    ws_sc.cell(row=26, column=sc_col, value=f"=IF(INPUTS!C7>0,{sc_col_letter}18*12/INPUTS!C7,0)")
    ws_sc.cell(row=26, column=sc_col).number_format = PCT_FORMAT

# Style result rows
for r in range(14, 27):
    style_row(ws_sc, r, 5, None, NORMAL_FONT)
    for c in [3, 4, 5]:
        ws_sc.cell(row=r, column=c).fill = RESULT_FILL
        ws_sc.cell(row=r, column=c).font = RESULT_FONT

# Highlight key rows
for r in [22, 23]:
    for c in [3, 4, 5]:
        ws_sc.cell(row=r, column=c).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws_sc.cell(row=r, column=c).font = Font(name="Calibri", bold=True, size=12, color="006100")

# Scenario comparison chart
chart_sc = BarChart()
chart_sc.type = "col"
chart_sc.title = "Сравнение сценариев — EBITDA и Чистая прибыль"
chart_sc.y_axis.title = "EUR"
chart_sc.style = 10
# EBITDA row
data_sc1 = Reference(ws_sc, min_col=3, min_row=22, max_col=5, max_row=22)
data_sc2 = Reference(ws_sc, min_col=3, min_row=23, max_col=5, max_row=23)
cats_sc = Reference(ws_sc, min_col=3, min_row=5, max_col=5, max_row=5)
chart_sc.add_data(data_sc1, from_rows=True, titles_from_data=False)
chart_sc.add_data(data_sc2, from_rows=True, titles_from_data=False)
chart_sc.set_categories(cats_sc)
chart_sc.series[0].title = openpyxl.chart.series.SeriesLabel(v="EBITDA")
chart_sc.series[1].title = openpyxl.chart.series.SeriesLabel(v="Чистая прибыль")
chart_sc.width = 22
chart_sc.height = 14
ws_sc.add_chart(chart_sc, "B28")


# ═══════════════════════════════════════════════════════════════════════════
# Final: Save
# ═══════════════════════════════════════════════════════════════════════════
output_path = "/home/user/VoyageRent/VoyageRent_Financial_Model.xlsx"
wb.save(output_path)
print(f"Financial model saved to: {output_path}")
